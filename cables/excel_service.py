from io import BytesIO
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from .calculations import APPROVED_AREAS_BY_BREAKER, APPROVED_BREAKERS_A
from .calculations import calculate_temperature_scenarios


BATCH_HEADERS = [
    "Project Reference", "Site Name", "Rack Name", "Engineer",
    "Calculation Date", "Load Current (A)", "Cable Length (m)",
    "Temperature Below 25C", "Temperature 25-30C",
    "Temperature 30-60C", "Temperature Above 60C",
    "Notes", "Approval Status",
]
APPROVAL_STATUSES = {"Draft", "Submitted", "Approved", "Rejected"}


def _clean_text(value):
    return "" if value is None else str(value).strip()


def _clean_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _clean_text(value)


def read_batch_workbook(source):
    """Validate the batch template and calculate every populated circuit row."""
    try:
        workbook = load_workbook(source, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("The uploaded file could not be read as an Excel workbook.") from exc
    if "Batch Inputs" not in workbook.sheetnames:
        raise ValueError("The workbook must contain a sheet named 'Batch Inputs'.")

    sheet = workbook["Batch Inputs"]
    actual_headers = [sheet.cell(4, column).value for column in range(1, len(BATCH_HEADERS) + 1)]
    if actual_headers != BATCH_HEADERS:
        raise ValueError("The Batch Inputs headers were changed. Download a fresh template and keep row 4 unchanged.")

    batch_rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=5, max_col=len(BATCH_HEADERS), values_only=True), start=5):
        if not any(value not in (None, "") for value in values):
            continue
        (
            project_reference, site_name, rack_name, engineer, calculation_date,
            load_raw, length_raw, temp_below_25, temp_25_30, temp_30_60,
            temp_above_60, notes, approval_raw,
        ) = values
        row = {
            "source_row": row_number,
            "project_reference": _clean_text(project_reference),
            "site_name": _clean_text(site_name),
            "rack_name": _clean_text(rack_name),
            "engineer": _clean_text(engineer),
            "calculation_date": _clean_date(calculation_date),
            "notes": _clean_text(notes),
            "approval_status": _clean_text(approval_raw) or "Draft",
            "load_current_a": load_raw,
            "length_m": length_raw,
            "temperature_below_25": temp_below_25,
            "temperature_25_30": temp_25_30,
            "temperature_30_60": temp_30_60,
            "temperature_above_60": temp_above_60,
        }
        errors = []
        for key, label in (
            ("project_reference", "Project Reference"), ("site_name", "Site Name"),
            ("rack_name", "Rack Name"), ("engineer", "Engineer"),
            ("calculation_date", "Calculation Date"),
        ):
            if not row[key]:
                errors.append(f"{label} is required")
        if row["approval_status"] not in APPROVAL_STATUSES:
            errors.append("Approval Status must be Draft, Submitted, Approved or Rejected")
        try:
            row["load_current_a"] = float(load_raw)
            if not 0 < row["load_current_a"] <= 80:
                errors.append("Load Current must be greater than 0 and no more than 80 A")
        except (TypeError, ValueError):
            errors.append("Load Current must be numeric")
        try:
            row["length_m"] = float(length_raw)
            if row["length_m"] <= 0:
                errors.append("Cable Length must be greater than 0 m")
        except (TypeError, ValueError):
            errors.append("Cable Length must be numeric")
        temperature_checks = (
            ("temperature_below_25", temp_below_25, lambda value: value < 25, "Temperature Below 25C must be below 25°C"),
            ("temperature_25_30", temp_25_30, lambda value: 25 <= value < 30, "Temperature 25-30C must be from 25°C to below 30°C"),
            ("temperature_30_60", temp_30_60, lambda value: 30 <= value <= 60, "Temperature 30-60C must be from 30°C to 60°C"),
            ("temperature_above_60", temp_above_60, lambda value: value > 60, "Temperature Above 60C must be above 60°C"),
        )
        for key, raw_value, condition, message in temperature_checks:
            try:
                row[key] = float(raw_value)
                if not condition(row[key]):
                    errors.append(message)
            except (TypeError, ValueError):
                errors.append(f"{message.split(' must')[0]} must be numeric")

        if errors:
            row.update({
                "required_breaker_a": None, "selected_breaker_a": None,
                "recommended_cable_mm2": None, "voltage_drop_v": None,
                "maximum_length_m": None, "result_status": "ERROR",
                "temperature_cases": [],
                "message": "; ".join(errors),
            })
        else:
            temperatures = {
                "below_25": row["temperature_below_25"],
                "between_25_30": row["temperature_25_30"],
                "between_30_60": row["temperature_30_60"],
                "above_60": row["temperature_above_60"],
            }
            selection, cases, recommendation = calculate_temperature_scenarios(
                row["load_current_a"], row["length_m"], temperatures, 3.0,
            )
            case_summaries = [{
                "key": case["key"], "label": case["label"],
                "temperature_c": case["temperature_c"], "resistivity": case["resistivity"],
                "recommended_cable_mm2": case["recommendation"]["size_mm2"] if case["recommendation"] else None,
                "voltage_drop_v": case["recommendation"]["voltage_drop"] if case["recommendation"] else None,
            } for case in cases]
            row.update({
                "required_breaker_a": selection["required_breaker_a"],
                "selected_breaker_a": selection["selected_breaker_a"],
                "recommended_cable_mm2": recommendation["size_mm2"] if recommendation else None,
                "voltage_drop_v": recommendation["voltage_drop"] if recommendation else None,
                "maximum_length_m": recommendation["maximum_length"] if recommendation else None,
                "temperature_cases": case_summaries,
                "result_status": "PASS" if recommendation else "NO MATCH",
                "message": "Overall cable passes all four conductor-temperature cases" if recommendation else "No approved cable satisfies all four temperature cases",
            })
        batch_rows.append(row)

    if not batch_rows:
        raise ValueError("No populated circuit rows were found in the Batch Inputs sheet.")
    if len(batch_rows) > 500:
        raise ValueError("A batch may contain no more than 500 circuit rows.")
    return batch_rows


def create_batch_results_workbook(batch_rows):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Batch Summary"
    red, dark_red = "E60000", "B80000"
    pass_count = sum(row["result_status"] == "PASS" for row in batch_rows)
    error_count = sum(row["result_status"] == "ERROR" for row in batch_rows)
    no_match_count = sum(row["result_status"] == "NO MATCH" for row in batch_rows)

    summary.merge_cells("A1:D2")
    summary["A1"] = "Vodafone DC Cable Batch Results — Phase 1"
    summary["A1"].fill = PatternFill("solid", fgColor=red)
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    summary["A1"].alignment = Alignment(vertical="center")
    summary["A4"] = "Summary item"
    summary["B4"] = "Count"
    for cell in summary[4]:
        cell.fill = PatternFill("solid", fgColor=dark_red)
        cell.font = Font(color="FFFFFF", bold=True)
    for label, value in (
        ("Total uploaded circuits", len(batch_rows)), ("Successful recommendations", pass_count),
        ("No approved cable match", no_match_count), ("Rows with input errors", error_count),
        ("Fixed voltage-drop limit (V)", 3), ("CB safety margin (%)", 25),
    ):
        summary.append([label, value])
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 20

    results = workbook.create_sheet("Batch Results")
    headers = BATCH_HEADERS + ["Required CB (A)", "Selected CB (A)"]
    for label in ("Below 25C", "25-30C", "30-60C", "Above 60C"):
        headers.extend([f"Resistivity at {label}", f"Cable at {label} (mm2)", f"Voltage Drop at {label} (V)"])
    headers += [
        "Overall Cable (mm2)", "Worst-case Voltage Drop (V)",
        "Worst-case Maximum Length (m)", "Result", "Validation Message",
    ]
    results.append(headers)
    for cell in results[1]:
        cell.fill = PatternFill("solid", fgColor=dark_red)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    results.freeze_panes = "A2"
    results.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(batch_rows) + 1}"
    for row in batch_rows:
        output_row = [
            row["project_reference"], row["site_name"], row["rack_name"], row["engineer"],
            row["calculation_date"], row["load_current_a"], row["length_m"],
            row["temperature_below_25"], row["temperature_25_30"],
            row["temperature_30_60"], row["temperature_above_60"],
            row["notes"], row["approval_status"], row["required_breaker_a"], row["selected_breaker_a"],
        ]
        for case in row.get("temperature_cases", []):
            output_row.extend([case["resistivity"], case["recommended_cable_mm2"], case["voltage_drop_v"]])
        while len(output_row) < len(BATCH_HEADERS) + 2 + 12:
            output_row.extend([None, None, None])
        output_row.extend([
            row["recommended_cable_mm2"], row["voltage_drop_v"], row["maximum_length_m"],
            row["result_status"], row["message"],
        ])
        results.append(output_row)
        fill = "E2F0D9" if row["result_status"] == "PASS" else "FCE4D6"
        for cell in results[results.max_row]:
            cell.fill = PatternFill("solid", fgColor=fill)
    for index, header in enumerate(headers, start=1):
        if header in {"Notes", "Validation Message"}:
            width = 42
        elif "Resistivity" in header or "Voltage Drop" in header:
            width = 23
        else:
            width = 18
        results.column_dimensions[get_column_letter(index)].width = width
    for cell in results[1]:
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_number in range(2, results.max_row + 1):
        for column in range(len(BATCH_HEADERS) + 3, len(headers) - 4):
            results.cell(row_number, column).number_format = "0.00000"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def create_results_workbook(selection, temperature_cases, recommendation):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Selection Summary"
    red, dark_red = "E60000", "BD0000"
    summary.merge_cells("A1:D2")
    summary["A1"] = "Vodafone DC Cable Selection — Phase 1"
    summary["A1"].fill = PatternFill("solid", fgColor=red)
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    summary["A1"].alignment = Alignment(vertical="center")
    summary["A4"] = "Design item"
    summary["B4"] = "Value"
    for cell in summary[4]:
        cell.fill = PatternFill("solid", fgColor=dark_red)
        cell.font = Font(color="FFFFFF", bold=True)
    values = [
        ("Load current (A)", selection["load_current_a"]),
        ("One-way length (m)", selection["length_m"]),
        ("Safety margin (%)", selection["safety_margin_percent"]),
        ("Required breaker (A)", selection["required_breaker_a"]),
        ("Selected approved breaker (A)", selection["selected_breaker_a"]),
        ("Maximum voltage drop (V)", selection["voltage_drop_limit_v"]),
        ("Base copper resistivity at 20C", selection["base_resistivity"]),
        ("Copper temperature coefficient", selection["temperature_coefficient"]),
        ("Overall cable passing all cases (mm2)", recommendation["size_mm2"] if recommendation else "No listed match"),
        ("Worst-case voltage drop (V)", recommendation["voltage_drop"] if recommendation else "—"),
        ("Worst-case conductor temperature (C)", recommendation["worst_case"]["temperature_c"] if recommendation else "—"),
    ]
    for label, value in values:
        summary.append([label, value])
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 25

    checks = workbook.create_sheet("Temperature Cases")
    checks.append(["Temperature case", "Conductor temperature (C)", "Corrected resistivity", "Recommended cable (mm2)", "Voltage drop (V)", "Maximum length (m)", "Status"])
    for cell in checks[1]:
        cell.fill = PatternFill("solid", fgColor=red)
        cell.font = Font(color="FFFFFF", bold=True)
    for case in temperature_cases:
        case_recommendation = case["recommendation"]
        checks.append([
            case["label"], case["temperature_c"], case["resistivity"],
            case_recommendation["size_mm2"] if case_recommendation else "No match",
            case_recommendation["voltage_drop"] if case_recommendation else None,
            case_recommendation["maximum_length"] if case_recommendation else None,
            "PASS" if case_recommendation else "NO MATCH",
        ])
    for row in checks.iter_rows(min_row=2):
        if row[6].value == "PASS":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E2F0D9")
    for index, width in enumerate([28, 25, 23, 26, 22, 22, 16], start=1):
        checks.column_dimensions[get_column_letter(index)].width = width
    for row_number in range(2, checks.max_row + 1):
        checks.cell(row_number, 3).number_format = "0.00000"
        checks.cell(row_number, 5).number_format = "0.000"
        checks.cell(row_number, 6).number_format = "0.00"

    candidates = workbook.create_sheet("All Candidate Checks")
    candidate_headers = ["Temperature case", "Temperature (C)", "Resistivity", "Cable area (mm2)", "Voltage drop (V)", "Maximum length (m)", "Result"]
    candidates.append(candidate_headers)
    for cell in candidates[1]:
        cell.fill = PatternFill("solid", fgColor=red)
        cell.font = Font(color="FFFFFF", bold=True)
    for case in temperature_cases:
        for row in case["results"]:
            candidates.append([
                case["label"], case["temperature_c"], case["resistivity"],
                row["size_mm2"], row["voltage_drop"], row["maximum_length"],
                "PASS" if row["overall_pass"] else "FAIL",
            ])
    candidates.freeze_panes = "A2"
    candidates.auto_filter.ref = f"A1:G{candidates.max_row}"
    for index, width in enumerate([28, 18, 20, 22, 20, 22, 14], start=1):
        candidates.column_dimensions[get_column_letter(index)].width = width

    reference = workbook.create_sheet("Approved Reference")
    reference.append(["CB rating (A)", "Maximum load (A)", "Approved cable areas (mm2)"])
    for cell in reference[1]:
        cell.fill = PatternFill("solid", fgColor=red)
        cell.font = Font(color="FFFFFF", bold=True)
    for breaker in APPROVED_BREAKERS_A:
        reference.append([breaker, breaker / 1.25, ", ".join(map(str, APPROVED_AREAS_BY_BREAKER[breaker]))])
    reference.column_dimensions["A"].width = 20
    reference.column_dimensions["B"].width = 22
    reference.column_dimensions["C"].width = 42

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
